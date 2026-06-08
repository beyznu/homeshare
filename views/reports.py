import tkinter as tk
import ttkbootstrap as ttk
from ttkbootstrap.constants import *
from tkinter import messagebox as msg
from tkinter import filedialog as fd
import openpyxl
import openpyxl.styles
import db_manager as dbm


class ReportsWindow(tk.Toplevel):

    def __init__(self, parent, user_id):
        super().__init__(parent)
        self.parent = parent
        self.user_id = user_id
        self.db = dbm.HomeShareDatabase()

        self.title("Rapor / Dışa Aktar")
        self.geometry("500x420")
        self.resizable(False, False)

        self.build_ui()
        self.load_preview()

        self.protocol("WM_DELETE_WINDOW", self.destroy)
        self._center_window()

    def build_ui(self):
        frm_info = ttk.LabelFrame(self, text="Excel Raporu", bootstyle="success")
        frm_info.pack(fill="x", padx=15, pady=(15, 0))

        ttk.Label(frm_info,
                  text="Tüm harcamalar, ödemeler ve bakiye özeti\nExcel dosyasına aktarılacak.",
                  font=("Helvetica", 10), bootstyle="success").pack(anchor="w", padx=10, pady=10)

        self.btn_export = ttk.Button(frm_info, text="Excel Olarak Kaydet",
                                     bootstyle="success", command=self.on_export)
        self.btn_export.pack(anchor="e", padx=10, pady=(0, 10))

        frm_preview = ttk.LabelFrame(self, text="Harcama Önizleme", bootstyle="success")
        frm_preview.pack(fill="both", expand=True, padx=15, pady=10)

        cols = ("date", "description", "category", "amount", "paid_by")
        self.tv = ttk.Treeview(frm_preview, columns=cols, show="headings",
                               height=10, selectmode="none", bootstyle="success")

        self.tv.heading("date",        text="Tarih",    anchor="center")
        self.tv.heading("description", text="Açıklama", anchor="w")
        self.tv.heading("category",    text="Kategori", anchor="w")
        self.tv.heading("amount",      text="Tutar",    anchor="center")
        self.tv.heading("paid_by",     text="Ödeyen",   anchor="w")

        self.tv.column("date",        anchor="center", width=80)
        self.tv.column("description", anchor="w",      width=130)
        self.tv.column("category",    anchor="w",      width=80)
        self.tv.column("amount",      anchor="center", width=75)
        self.tv.column("paid_by",     anchor="w",      width=100)

        scroll = ttk.Scrollbar(frm_preview, orient="vertical", command=self.tv.yview,
                               bootstyle="success-round")
        self.tv.configure(yscrollcommand=scroll.set)
        scroll.pack(side="right", fill="y")
        self.tv.pack(fill="both", expand=True, padx=(8, 0), pady=8)

    def load_preview(self):
        for item in self.tv.get_children():
            self.tv.delete(item)

        rows = self.db.get_expenses(self.user_id)

        if not rows:
            self.tv.insert("", "end", values=("—", "Henüz harcama yok", "—", "—", "—"))
            self.btn_export.configure(state="disabled")
            return

        for row in rows:
            self.tv.insert("", "end", values=(
                row[3], row[1], row[5] or "—", f"{row[2]:.2f} ₺", row[4]
            ))

    def on_export(self):
        expenses = self.db.get_expenses(self.user_id)
        if not expenses:
            msg.showerror("Hata", "Dışa aktarılacak harcama verisi yok.", parent=self)
            return

        file_filter = (("Excel dosyası", "*.xlsx"), ("Tüm dosyalar", "*.*"))
        file_path = fd.asksaveasfilename(title="Raporu Kaydet", filetypes=file_filter,
                                         defaultextension=".xlsx",
                                         initialfile="homeshare_rapor")
        if not file_path:
            return

        wb = openpyxl.Workbook()

        ws_exp = wb.active
        ws_exp.title = "Harcamalar"
        ws_exp.append(["Tarih", "Açıklama", "Kategori", "Tutar (₺)", "Ödeyen"])

        for cell in ws_exp["1"]:
            cell.font = openpyxl.styles.Font(bold=True)

        for row in expenses:
            ws_exp.append([row[3], row[1], row[5] or "—", row[2], row[4]])

        last_data_row = ws_exp.max_row
        ws_exp.append([])
        ws_exp.append(["Toplam Harcama Sayısı:", f"=COUNTA(A2:A{last_data_row})"])
        ws_exp.append(["Toplam Tutar (₺):",      f"=SUM(D2:D{last_data_row})"])
        ws_exp.append(["Ortalama Tutar (₺):",    f"=AVERAGE(D2:D{last_data_row})"])

        ws_exp.column_dimensions["A"].width = 12
        ws_exp.column_dimensions["B"].width = 30
        ws_exp.column_dimensions["C"].width = 15
        ws_exp.column_dimensions["D"].width = 12
        ws_exp.column_dimensions["E"].width = 15

        ws_bal = wb.create_sheet("Bakiye Özeti")
        ws_bal.append(["Kişi", "Toplam Ödedi (₺)", "Toplam Payı (₺)", "Bakiye (₺)"])
        for cell in ws_bal["1"]:
            cell.font = openpyxl.styles.Font(bold=True)
        for r_id, name, paid, share, balance in self.db.get_balance_summary(self.user_id):
            ws_bal.append([name, paid, share, balance])
        ws_bal.column_dimensions["A"].width = 20
        ws_bal.column_dimensions["B"].width = 18
        ws_bal.column_dimensions["C"].width = 18
        ws_bal.column_dimensions["D"].width = 15

        ws_pay = wb.create_sheet("Ödemeler")
        ws_pay.append(["Tarih", "Ödeyen", "Alan", "Tutar (₺)", "Not"])
        for cell in ws_pay["1"]:
            cell.font = openpyxl.styles.Font(bold=True)
        for row in self.db.get_payments(self.user_id):
            ws_pay.append([row[2], row[4], row[5], row[1], row[3] or ""])
        ws_pay.column_dimensions["A"].width = 12
        ws_pay.column_dimensions["B"].width = 15
        ws_pay.column_dimensions["C"].width = 15
        ws_pay.column_dimensions["D"].width = 12
        ws_pay.column_dimensions["E"].width = 25

        wb.save(file_path)
        msg.showinfo("Başarılı", f"Rapor başarıyla kaydedildi:\n{file_path}", parent=self)

    def _center_window(self):
        self.update_idletasks()
        w = self.winfo_width()
        h = self.winfo_height()
        x = (self.winfo_screenwidth()  // 2) - (w // 2)
        y = (self.winfo_screenheight() // 2) - (h // 2)
        self.geometry(f"{w}x{h}+{x}+{y}")